import openpyxl
import json
import os

def export_analysis():
    file_path = '/Users/kristiyankirov/Downloads/IBEX_2025_vs_2026_Analysis.xlsx'
    output_path = 'public/data/ibex-analysis.json'
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    print(f"Reading {file_path} with openpyxl...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = wb.sheetnames
        
        target_sheets = [
            'Comparison_2025_vs_2026', 
            'Monthly_Summary', 
            'Battery_Summary', 
            'Daily_Battery', 
            'Hourly_Profile', 
            'Executive_Insights'
        ]
        
        analysis_data = {}
        
        for sheet_name in target_sheets:
            if sheet_name in sheets:
                print(f"Extracting {sheet_name}...")
                ws = wb[sheet_name]
                data = []
                
                # Special handling for Executive_Insights which might not be a standard table
                if sheet_name == 'Executive_Insights':
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell and isinstance(cell, str) and len(cell) > 20:
                                data.append({"insight": cell})
                else:
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: continue
                    
                    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
                    
                    for row in rows[1:]:
                        if all(v is None for v in row): continue
                        row_data = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                header = headers[i]
                                if hasattr(value, 'isoformat'):
                                    value = value.isoformat()
                                row_data[header] = value if value is not None else ""
                        data.append(row_data)
                
                analysis_data[sheet_name] = data
            else:
                print(f"Warning: Sheet {sheet_name} not found.")

        # Save to JSON
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w') as f:
            json.dump(analysis_data, f, indent=2)
            
        print(f"Successfully exported analysis to {output_path}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    export_analysis()
